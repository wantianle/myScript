import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- 配置区 ---
CONFIG = {
    "SETTLEMENT": ["谢畅", "余达柔"],
    "HYBRID": ["林琳", "曾丽敏", "李建东"],
    "M3_05": ["聂玉芬", "管玉霞"],
    "QIWEI_FIXED": ["闫雪001", "徐璐"],  # 固定企微客服的人员
    "COLOR_HEADER": "DDEBF7",
    "COLOR_SUBTOTAL": "FFF2CC",
    "COLOR_GRANDTOTAL": "E2EFDA",
}


def parse_seconds(t_val):
    if pd.isna(t_val) or str(t_val).strip() in ["", "--", "0"]:
        return 0
    parts = str(t_val).strip().split(":")
    return (
        int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        if len(parts) == 3
        else 0
    )


def format_duration(seconds):
    return f"{int(seconds // 3600)}小时{int((seconds % 3600) // 60)}分钟"


def get_cleansed_df(path, is_ext):
    df = pd.read_excel(path)
    df["时间"] = pd.to_datetime(df["会话结束时间"]).dt.strftime("%Y-%m-%d")
    df["s1_sec"] = df["客服首次响应时长"].apply(parse_seconds)
    df["sa_sec"] = df["客服平均响应时长"].apply(parse_seconds)
    df["sr_sec"] = df["人工接待时长"].apply(parse_seconds)
    df["is_valid"] = (df["是否有效会话"] == "有效会话").astype(int)
    df["is_invited"] = (df["客服是否邀评"] == "是").astype(int)
    df["is_eval"] = (df["满意度"] != "未评价").astype(int)
    df["is_good"] = df["满意度"].isin(["非常满意", "满意"]).astype(int)

    def assign_type(row):
        n, tag = row["接待客服"], str(row["客户标签"])
        if not is_ext:
            return "对内客服"
        # 1. 优先判定固定企微人员
        if n in CONFIG["QIWEI_FIXED"]:
            return "企微客服"
        # 2. 判定结算人员
        if n in CONFIG["SETTLEMENT"]:
            return "结算客服"
        # 3. 判定交叉岗（林琳等三个人）
        if n in CONFIG["HYBRID"]:
            return "M3" if "M3" in tag else "企微客服"
        # 4. 其余默认 M3
        return "M3"

    df["类型"] = df.apply(assign_type, axis=1)
    return df


def aggregate_with_totals(df, is_ext):
    group_keys = ["时间", "类型", "接待客服"]
    base_agg = (
        df.groupby(group_keys)
        .agg(
            {
                "会话ID": "count",
                "is_valid": "sum",
                "is_invited": "sum",
                "is_eval": "sum",
                "is_good": "sum",
                "s1_sec": "sum",
                "sa_sec": "sum",
                "sr_sec": "sum",
            }
        )
        .reset_index()
    )

    output_list = []
    for day, day_df in base_agg.groupby("时间"):
        all_names = day_df["接待客服"].unique()
        day_total_fte = sum(
            [
                0.5 if (n in CONFIG["HYBRID"] or n in CONFIG["M3_05"]) else 1.0
                for n in all_names
            ]
        )

        day_m = {
            k: 0.0 for k in ["cnt", "val", "fte", "inv", "ev", "gd", "s1", "sa", "sr"]
        }
        types = ["企微客服", "M3", "结算客服"] if is_ext else ["对内客服"]

        for t in types:
            t_df = day_df[day_df["类型"] == t]
            if t_df.empty:
                continue

            # 只要交叉岗在这个分类露面，就给该分类加0.5人力
            names_in_type = t_df["接待客服"].unique()
            type_fte = sum(
                [
                    0.5 if (n in CONFIG["HYBRID"] or n in CONFIG["M3_05"]) else 1.0
                    for n in names_in_type
                ]
            )

            sub_m = {
                k: 0.0 for k in ["cnt", "val", "fte", "inv", "ev", "gd", "s1", "sa", "sr"]
            }
            for _, r in t_df.iterrows():
                name = r["接待客服"]
                p_fte = (
                    0.5
                    if (name in CONFIG["HYBRID"] or name in CONFIG["M3_05"])
                    else 1.0
                )
                row = {
                    "时间": day,
                    "类型": t,
                    "人员": name,
                    "cnt": r["会话ID"],
                    "val": r["is_valid"],
                    "fte": p_fte,
                    "inv": r["is_invited"],
                    "ev": r["is_eval"],
                    "gd": r["is_good"],
                    "s1": r["s1_sec"],
                    "sa": r["sa_sec"],
                    "sr": r["sr_sec"],
                }
                output_list.append(row)
                for k in ["cnt", "val", "inv", "ev", "gd", "s1", "sa", "sr"]:
                    sub_m[k] += row[k]

            sub_m["fte"] = type_fte
            if is_ext:
                output_list.append({"时间": day, "类型": t, "人员": "合计", **sub_m})
            for k in ["cnt", "val", "inv", "ev", "gd", "s1", "sa", "sr"]:
                day_m[k] += sub_m[k]

        day_m["fte"] = day_total_fte
        output_list.append(
            {
                "时间": day,
                "类型": "当日总计" if is_ext else "对内客服",
                "人员": "合计",
                **day_m,
            }
        )

    res = pd.DataFrame(output_list)

    # 逻辑修正：分母为0显示斜杠
    res["相对满意度(%)"] = res.apply(
        lambda x: x["gd"] / x["ev"] if x["ev"] > 0 else "/", axis=1
    )
    res["平均首次响应时长（s）"] = res.apply(
        lambda x: round(x["s1"] / x["val"], 1) if x["val"] > 0 else "/", axis=1
    )
    res["平均会话响应时长（s）"] = res.apply(
        lambda x: round(x["sa"] / x["val"], 1) if x["val"] > 0 else "/", axis=1
    )
    res["人工接待时长总计（00:00）"] = res["sr"].apply(format_duration)

    final_cols_map = {
        "时间": "时间",
        "类型": "类型",
        "人员": "人员",
        "cnt": "总会话量",
        "val": "有效会话",
        "fte": "对内FTE",
        "inv": "邀评数",
        "ev": "参评数",
        "相对满意度(%)": "相对满意度(%)",
        "平均首次响应时长（s）": "平均首次响应时长（s）",
        "平均会话响应时长（s）": "平均会话响应时长（s）",
        "人工接待时长总计（00:00）": "人工接待时长总计（00:00）",
    }
    if not is_ext:
        final_cols_map.pop("类型")
    return res[list(final_cols_map.keys())].rename(columns=final_cols_map)


# apply_excel_formatting 函数同之前修正版
def apply_excel_formatting(ws, is_ext):
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    max_row, max_col = ws.max_row, ws.max_column
    for cell in ws[1]:
        cell.fill = PatternFill(
            start_color=CONFIG["COLOR_HEADER"],
            end_color=CONFIG["COLOR_HEADER"],
            fill_type="solid",
        )
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    merge_cols = [1, 2] if is_ext else [1]
    for col_idx in merge_cols:
        start_r = 2
        for r in range(3, max_row + 2):
            curr_val = ws.cell(r, col_idx).value
            prev_val = ws.cell(start_r, col_idx).value
            if curr_val in ["当日总计", "合计"] or r > max_row or curr_val != prev_val:
                if r - start_r > 1:
                    ws.merge_cells(
                        start_row=start_r,
                        start_column=col_idx,
                        end_row=r - 1,
                        end_column=col_idx,
                    )
                ws.cell(start_r, col_idx).alignment = Alignment(
                    horizontal="center", vertical="top"
                )
                start_r = r

    person_col_idx, type_col_idx = (3, 2) if is_ext else (2, 1)
    for r in range(2, max_row + 1):
        p_val, t_val = (
            str(ws.cell(r, person_col_idx).value),
            str(ws.cell(r, type_col_idx).value),
        )
        fill = None
        if t_val == "当日总计" or (not is_ext and p_val == "合计"):
            fill = PatternFill(
                start_color=CONFIG["COLOR_GRANDTOTAL"],
                end_color=CONFIG["COLOR_GRANDTOTAL"],
                fill_type="solid",
            )
        elif p_val == "合计":
            fill = PatternFill(
                start_color=CONFIG["COLOR_SUBTOTAL"],
                end_color=CONFIG["COLOR_SUBTOTAL"],
                fill_type="solid",
            )

        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if fill:
                cell.fill = fill
                cell.font = Font(bold=True)
            if cell.value == "/":
                cell.alignment = Alignment(horizontal="center")
            else:
                if "满意度" in str(ws.cell(1, c).value) and isinstance(
                    cell.value, (int, float)
                ):
                    cell.number_format = "0.0%"
                cell.alignment = (
                    Alignment(horizontal="right")
                    if isinstance(cell.value, (int, float))
                    else Alignment(horizontal="center")
                )
    for i in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


def run_task(ext_file, int_file):
    res_ext = aggregate_with_totals(get_cleansed_df(ext_file, True), True)
    res_int = aggregate_with_totals(get_cleansed_df(int_file, False), False)
    output = "客服运营分析报告.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        res_ext.to_excel(writer, sheet_name="对外客服", index=False)
        apply_excel_formatting(writer.sheets["对外客服"], True)
        res_int.to_excel(writer, sheet_name="对内客服", index=False)
        apply_excel_formatting(writer.sheets["对内客服"], False)
    print(f"✅ 处理成功：{output}")


run_task("对外的表.xlsx", "对内的表.xlsx")
